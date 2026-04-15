import os
import json
import logging
import copy
import pandas as pd
from openpyxl.styles import Font

from .excel_utils import sanitize_for_excel

logger = logging.getLogger(__name__)

COLUMNS = [
    "id", "repository", "file_path", "line_number",
    "secret_type", "secret_value", "commit", "found_by",
]


def _build_commit_url(repo_url, commit_hash, file_path, line_number=""):
    """Build a browser URL pointing to a specific file at a given commit.

    Supports GitHub and GitLab URL patterns.  Returns an empty string
    when there is not enough information to construct a valid link.

    .. note::

       If the file was deleted or renamed after this commit the link may
       404.  See the README for a quick workaround (replace ``/blob/``
       with ``/commit/`` and drop the file path).
    """
    if not repo_url or not commit_hash:
        return ""
    # GitLab uses /-/blob/…, everything else (GitHub, Gitea, …) uses /blob/…
    if "gitlab" in repo_url.lower():
        url = f"{repo_url}/-/blob/{commit_hash}/{file_path}"
    else:
        url = f"{repo_url}/blob/{commit_hash}/{file_path}"
    if line_number:
        url += f"#L{line_number}"
    return url



class Reporter:
    def __init__(self, output_dir, repo_name=""):
        self.output_dir = output_dir
        self.repo_name = repo_name

    def _prefixed(self, filename):
        """Prepend repo_name_ to filename when a repo name is set."""
        if self.repo_name:
            return f"{self.repo_name}_{filename}"
        return filename

    def generate_json(self, data):
        json_path = os.path.join(self.output_dir, self._prefixed("aggregated_secrets.json"))
        logger.info(f"Generating aggregated JSON report at {json_path}")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
        logger.info(f"Wrote {len(data)} findings to {json_path}")
        return json_path

    def _prepare_excel_data(self, data):
        """Convert found_by lists to comma-separated strings and
        rename commit_hash → commit (short hash) for Excel readability.

        The internal ``repo_url`` field is stripped — it is only used
        for building hyperlinks, not displayed as a column.
        """
        excel_data = []
        for item in data:
            row = copy.deepcopy(item)
            if isinstance(row.get("found_by"), list):
                row["found_by"] = ", ".join(row["found_by"])
            # Rename commit_hash → commit (full hash)
            commit_hash = row.pop("commit_hash", "")
            row.pop("repo_url", None)
            row["commit"] = commit_hash
            excel_data.append(row)
        return excel_data

    @staticmethod
    def _get_commit_urls(data):
        """Return a list of commit URLs aligned with *data* rows."""
        urls = []
        for item in data:
            url = _build_commit_url(
                item.get("repo_url", ""),
                item.get("commit_hash", ""),
                item.get("file_path", ""),
                item.get("line_number", ""),
            )
            urls.append(url)
        return urls

    @staticmethod
    def _add_commit_hyperlinks(ws, urls):
        """Post-process an openpyxl worksheet to turn the *commit* column
        into clickable hyperlinks."""
        # Locate the commit column by header name
        commit_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "commit":
                commit_col = col
                break
        if commit_col is None:
            return

        link_font = Font(color="0563C1", underline="single")
        for row_idx, url in enumerate(urls, start=2):
            if url:
                cell = ws.cell(row=row_idx, column=commit_col)
                cell.hyperlink = url
                cell.font = link_font

    def generate_excel(self, data):
        excel_path = os.path.join(self.output_dir, self._prefixed("secrets_report.xlsx"))
        logger.info(f"Generating Excel report at {excel_path}")

        if not data:
            logger.warning("No data to write to Excel.")
            df = pd.DataFrame(columns=COLUMNS)
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="General", index=False)
                ws = writer.sheets["General"]
                ws.auto_filter.ref = ws.dimensions
            return excel_path

        excel_data = self._prepare_excel_data(data)
        commit_urls_all = self._get_commit_urls(data)
        df_all = pd.DataFrame(excel_data)[COLUMNS]

        # Determine all unique tools
        tools = set()
        for item in data:
            tools.update(item["found_by"])

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # General tab — all deduplicated findings
            sanitize_for_excel(df_all).to_excel(writer, sheet_name="General", index=False)
            self._add_commit_hyperlinks(writer.sheets["General"], commit_urls_all)

            # Per-tool tabs
            for tool in sorted(tools):
                tool_items = [item for item in data if tool in item["found_by"]]
                tool_excel = self._prepare_excel_data(tool_items)
                tool_urls = self._get_commit_urls(tool_items)
                df_tool = pd.DataFrame(tool_excel)
                if not df_tool.empty:
                    df_tool = df_tool[COLUMNS]
                    sheet_name = str(tool).capitalize()[:31]
                    sanitize_for_excel(df_tool).to_excel(writer, sheet_name=sheet_name, index=False)
                    self._add_commit_hyperlinks(writer.sheets[sheet_name], tool_urls)

            # Add auto-filters to every sheet
            for ws in writer.sheets.values():
                ws.auto_filter.ref = ws.dimensions

        logger.info(f"Wrote Excel report with {len(tools)} tool tabs to {excel_path}")
        return excel_path
