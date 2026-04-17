"""Shared Excel utilities.

Used by both ``core.reporter`` (scan results) and
``core.ai.triage_reporter`` (AI triage results) to prepare DataFrames
for openpyxl — sanitization, URL building, and commit hyperlinks.
"""

import re

import pandas as pd
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# Control characters that are illegal in XML / Excel cells.
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Truncate abnormally large values that would break Excel (limit: 32 767 chars).
SECRET_VALUE_LIMIT = 5000


def sanitize_for_excel(df):
    """Make a DataFrame safe for openpyxl.

    1. Strip control characters that are illegal in XML/Excel.
    2. Truncate ``secret_value`` when abnormally large (>5 000 chars).

    Full values are always available in the JSON report.
    """
    df = df.copy()
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].apply(
            lambda v: ILLEGAL_CHARS_RE.sub("", v) if isinstance(v, str) else v
        )
    if "secret_value" in df.columns:
        df["secret_value"] = df["secret_value"].apply(
            lambda v: v[:SECRET_VALUE_LIMIT] + " [truncated]"
            if isinstance(v, str) and len(v) > SECRET_VALUE_LIMIT
            else v
        )
    return df


def _normalize_repo_url(repo_url):
    """Convert SSH or HTTPS git URLs to a browsable HTTPS base URL.

    Examples::

        git@github.com:org/repo.git  ->  https://github.com/org/repo
        https://github.com/org/repo.git  ->  https://github.com/org/repo
        https://github.com/org/repo  ->  https://github.com/org/repo
    """
    url = repo_url.strip()
    # SSH format: git@host:org/repo.git
    m = re.match(r"^git@([^:]+):(.+?)(?:\.git)?$", url)
    if m:
        return f"https://{m.group(1)}/{m.group(2)}"
    # HTTPS with .git suffix
    if url.endswith(".git"):
        url = url[:-4]
    return url


def build_commit_url(repo_url, commit_hash, file_path, line_number=""):
    """Build a browser URL pointing to a specific file at a given commit.

    Supports GitHub and GitLab URL patterns.  Automatically normalizes
    SSH URLs (``git@host:org/repo.git``) to HTTPS.  Returns an empty
    string when there is not enough information to construct a valid link.
    """
    if not repo_url or not commit_hash:
        return ""
    base = _normalize_repo_url(repo_url)
    if "gitlab" in base.lower():
        url = f"{base}/-/blob/{commit_hash}/{file_path}"
    else:
        url = f"{base}/blob/{commit_hash}/{file_path}"
    if line_number:
        url += f"#L{line_number}"
    return url


def add_commit_hyperlinks(ws, repo_url, commit_data):
    """Turn the ``commit`` column into clickable hyperlinks.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
    repo_url : str
        Remote origin URL (e.g. ``https://github.com/org/repo``).
    commit_data : list[dict] | pandas.DataFrame
        Rows with ``commit``, ``file_path``, ``line_number`` fields,
        aligned with the worksheet data rows (starting at row 2).
    """
    if not repo_url:
        return

    commit_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == "commit":
            commit_col = col_idx
            break
    if commit_col is None:
        return

    link_font = Font(color="0563C1", underline="single")

    if isinstance(commit_data, pd.DataFrame):
        rows = (row for _, row in commit_data.iterrows())
    else:
        rows = commit_data

    for row_idx, row in enumerate(rows, start=2):
        if isinstance(row, dict):
            commit = str(row.get("commit", row.get("commit_hash", "")))
            fp = str(row.get("file_path", ""))
            ln = str(row.get("line_number", ""))
        else:
            commit = str(row.get("commit", row.get("commit_hash", "")))
            fp = str(row.get("file_path", ""))
            ln = str(row.get("line_number", ""))

        url = build_commit_url(repo_url, commit, fp, ln)
        if url:
            cell = ws.cell(row=row_idx, column=commit_col)
            link = Hyperlink(ref=cell.coordinate, target=url)
            cell._hyperlink = link
            ws._hyperlinks.append(link)
            cell.font = link_font
