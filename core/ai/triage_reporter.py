"""Convert triage-results.json to an Excel workbook.

Adds triage-specific formatting (severity color-coding, TP/FP classification,
composite vulns tab) on top of shared Excel sanitization from
``core.excel_utils``.

Usage::

    from core.ai.triage_reporter import convert
    convert("path/to/triage-results.json")

Or from the CLI::

    python3 -m core.ai.triage_reporter <triage-results.json> [output.xlsx]
"""

import json
import logging
import os

import pandas as pd
from openpyxl.styles import Font, PatternFill

from core.excel_utils import sanitize_for_excel, SECRET_VALUE_LIMIT

logger = logging.getLogger(__name__)

# ── constants ────────────────────────────────────────────────────────────────

# Severity ordering for sorting (highest first).
_SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}

_SEVERITY_FILL = {
    "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "HIGH": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "LOW": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
}
_FP_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

FINDING_COLUMNS = [
    "id", "classification", "severity", "confidence", "category",
    "secret_value", "file_path", "line_number", "commit",
    "on_disk", "environment", "remediation", "effort",
    "detected_by", "fp_reason", "omnileak_ids",
]

COMPOSITE_COLUMNS = [
    "id", "severity", "description", "related_finding_ids", "files_involved",
]

# ── helpers ──────────────────────────────────────────────────────────────────


def _flatten_lists(df):
    """Convert list columns to comma-separated strings for Excel."""
    for col in ("detected_by", "omnileak_ids", "related_finding_ids", "files_involved"):
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: ", ".join(str(x) for x in v) if isinstance(v, list) else v
            )
    return df


def _sort_findings(df):
    """Sort by classification (TP first) then severity (CRITICAL first)."""
    df = df.copy()
    df["_cls_order"] = df["classification"].map(
        {"TRUE_POSITIVE": 0, "FALSE_POSITIVE": 1}
    ).fillna(2)
    df["_sev_order"] = df["severity"].map(_SEVERITY_ORDER).fillna(99)
    df = df.sort_values(["_cls_order", "_sev_order", "id"]).drop(
        columns=["_cls_order", "_sev_order"]
    )
    return df.reset_index(drop=True)


def _apply_colors(ws):
    """Color-code severity and classification cells."""
    sev_col = None
    cls_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header == "severity":
            sev_col = col_idx
        elif header == "classification":
            cls_col = col_idx

    for row_idx in range(2, ws.max_row + 1):
        if sev_col:
            val = ws.cell(row=row_idx, column=sev_col).value
            fill = _SEVERITY_FILL.get(val)
            if fill:
                ws.cell(row=row_idx, column=sev_col).fill = fill
        if cls_col:
            val = ws.cell(row=row_idx, column=cls_col).value
            if val == "FALSE_POSITIVE":
                ws.cell(row=row_idx, column=cls_col).fill = _FP_FILL


def _write_sheet(writer, df, sheet_name, columns):
    """Write a DataFrame to a named sheet with sanitization and formatting."""
    cols = [c for c in columns if c in df.columns]
    out = sanitize_for_excel(_flatten_lists(df[cols].copy()))
    out.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    ws.auto_filter.ref = ws.dimensions
    _apply_colors(ws)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)


# ── public API ───────────────────────────────────────────────────────────────


def convert(json_path, excel_path=None):
    """Read a ``triage-results.json`` and write ``triage-results.xlsx``.

    Parameters
    ----------
    json_path : str
        Path to the triage JSON produced by the scan-secrets skill.
    excel_path : str, optional
        Destination ``.xlsx`` path.  Defaults to ``triage-results.xlsx``
        in the same directory as *json_path*.

    Returns
    -------
    str
        The path of the written Excel file.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if excel_path is None:
        excel_path = os.path.join(
            os.path.dirname(json_path), "triage-results.xlsx"
        )

    findings = data.get("findings", [])
    composites = data.get("composite_vulnerabilities", [])
    meta = data.get("meta", {})

    df_all = pd.DataFrame(findings) if findings else pd.DataFrame(columns=FINDING_COLUMNS)
    df_all = _sort_findings(df_all)

    df_tp = df_all[df_all["classification"] == "TRUE_POSITIVE"].copy()
    df_fp = df_all[df_all["classification"] == "FALSE_POSITIVE"].copy()

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_rows = [
            {"Metric": "Repository", "Value": meta.get("repo", "")},
            {"Metric": "Scan Date", "Value": meta.get("scan_date", "")},
            {"Metric": "Last Commit", "Value": meta.get("last_commit", "")},
            {"Metric": "Mode", "Value": meta.get("mode", "")},
            {"Metric": "Risk Score", "Value": meta.get("risk_score", "")},
            {"Metric": "Total Raw Findings", "Value": meta.get("total_raw_findings", "")},
            {"Metric": "True Positives", "Value": meta.get("true_positives", "")},
            {"Metric": "False Positives Filtered", "Value": meta.get("false_positives_filtered", "")},
            {"Metric": "AI-Only Findings", "Value": meta.get("ai_only_findings", "")},
            {"Metric": "Deep Analysis", "Value": str(meta.get("deep_analysis_performed", ""))},
        ]
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws_summary = writer.sheets["Summary"]
        for col_idx in range(1, ws_summary.max_column + 1):
            ws_summary.cell(row=1, column=col_idx).font = Font(bold=True)

        _write_sheet(writer, df_all, "All Findings", FINDING_COLUMNS)

        if not df_tp.empty:
            _write_sheet(writer, df_tp, "True Positives", FINDING_COLUMNS)

        if not df_fp.empty:
            _write_sheet(writer, df_fp, "False Positives", FINDING_COLUMNS)

        if composites:
            df_comp = pd.DataFrame(composites)
            _write_sheet(writer, df_comp, "Composite Vulns", COMPOSITE_COLUMNS)

    logger.info(
        "Wrote triage Excel: %s (%d TPs, %d FPs, %d composites)",
        excel_path, len(df_tp), len(df_fp), len(composites),
    )
    return excel_path


# ── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print(f"Usage: python3 -m core.ai.triage_reporter <triage-results.json> [output.xlsx]")
        sys.exit(1)

    json_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.isfile(json_path):
        print(f"Error: {json_path} not found")
        sys.exit(1)

    result = convert(json_path, excel_path)
    print(f"Wrote {result}")
