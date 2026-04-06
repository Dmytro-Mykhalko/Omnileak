import os
import json
import pytest
import pandas as pd
from openpyxl import load_workbook
from core.reporter import Reporter, _build_commit_url


SAMPLE_DATA = [
    {
        "id": "aaa",
        "repository": "repo",
        "file_path": "config.yml",
        "line_number": 10,
        "secret_type": "AWS_KEY",
        "secret_value": "AKIA123",
        "commit_hash": "abc1234deadbeef",
        "repo_url": "https://example.com/org/repo",
        "found_by": ["gitleaks", "trufflehog"],
    },
    {
        "id": "bbb",
        "repository": "repo",
        "file_path": "env.sh",
        "line_number": 2,
        "secret_type": "DB_PASS",
        "secret_value": "s3cr3t",
        "commit_hash": "def5678beefdead",
        "repo_url": "https://example.com/org/repo",
        "found_by": ["detect-secrets"],
    },
]


class TestBuildCommitUrl:
    def test_github_style(self):
        url = _build_commit_url(
            "https://example.com/org/repo", "abc123", "src/app.py", 42
        )
        assert url == "https://example.com/org/repo/blob/abc123/src/app.py#L42"

    def test_gitlab_style(self):
        url = _build_commit_url(
            "https://gitlab.example.com/org/repo", "abc123", "src/app.py", 5
        )
        assert url == "https://gitlab.example.com/org/repo/-/blob/abc123/src/app.py#L5"

    def test_no_line_number(self):
        url = _build_commit_url(
            "https://example.com/org/repo", "abc123", "file.txt"
        )
        assert url == "https://example.com/org/repo/blob/abc123/file.txt"

    def test_empty_repo_url(self):
        assert _build_commit_url("", "abc123", "file.txt") == ""

    def test_empty_commit(self):
        assert _build_commit_url("https://example.com/org/repo", "", "f.txt") == ""


class TestJsonReport:
    def test_generates_json(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_json(SAMPLE_DATA)
        assert os.path.exists(path)
        assert os.path.basename(path) == "myrepo_aggregated_secrets.json"
        with open(path) as f:
            data = json.load(f)
        assert len(data) == 2
        assert data[0]["id"] == "aaa"

    def test_empty_json(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_json([])
        with open(path) as f:
            data = json.load(f)
        assert data == []

    def test_json_without_repo_name(self, tmp_path):
        reporter = Reporter(str(tmp_path))
        path = reporter.generate_json(SAMPLE_DATA)
        assert os.path.basename(path) == "aggregated_secrets.json"


class TestExcelReport:
    def test_generates_excel_with_tabs(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        assert os.path.exists(path)
        assert os.path.basename(path) == "myrepo_secrets_report.xlsx"

        xl = pd.ExcelFile(path)
        sheet_names = xl.sheet_names
        assert "General" in sheet_names
        assert "Gitleaks" in sheet_names
        assert "Trufflehog" in sheet_names
        assert "Detect-secrets" in sheet_names

    def test_general_tab_has_all_rows(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        df = pd.read_excel(path, sheet_name="General")
        assert len(df) == 2

    def test_commit_column_replaces_commit_hash(self, tmp_path):
        """Column should be named 'commit' with full hashes, not 'commit_hash'."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        df = pd.read_excel(path, sheet_name="General")
        assert "commit" in df.columns
        assert "commit_hash" not in df.columns
        assert df.iloc[0]["commit"] == "abc1234deadbeef"
        assert df.iloc[1]["commit"] == "def5678beefdead"

    def test_commit_column_has_hyperlinks(self, tmp_path):
        """Commit cells should contain clickable hyperlinks to the commit page."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        wb = load_workbook(path)
        ws = wb["General"]

        # Find commit column
        commit_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "commit":
                commit_col = col
                break
        assert commit_col is not None

        # Row 2 = first data row
        cell = ws.cell(row=2, column=commit_col)
        assert cell.hyperlink is not None
        expected = "https://example.com/org/repo/blob/abc1234deadbeef/config.yml#L10"
        assert cell.hyperlink.target == expected

    def test_commit_hyperlink_absent_without_repo_url(self, tmp_path):
        """When repo_url is empty, commit should be plain text (no link)."""
        data_no_url = [
            {
                "id": "ccc",
                "repository": "repo",
                "file_path": "f.py",
                "line_number": 1,
                "secret_type": "T",
                "secret_value": "val",
                "commit_hash": "aaa1111",
                "repo_url": "",
                "found_by": ["gitleaks"],
            }
        ]
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(data_no_url)
        wb = load_workbook(path)
        ws = wb["General"]

        commit_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "commit":
                commit_col = col
                break
        cell = ws.cell(row=2, column=commit_col)
        assert cell.hyperlink is None
        assert cell.value == "aaa1111"

    def test_commit_hyperlink_on_tool_tab(self, tmp_path):
        """Per-tool sheets should also have commit hyperlinks."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        wb = load_workbook(path)
        ws = wb["Gitleaks"]

        commit_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "commit":
                commit_col = col
                break
        cell = ws.cell(row=2, column=commit_col)
        assert cell.hyperlink is not None

    def test_found_by_is_string_in_excel(self, tmp_path):
        """found_by should be a comma-separated string, not a Python list."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        df = pd.read_excel(path, sheet_name="General")
        assert df.iloc[0]["found_by"] == "gitleaks, trufflehog"
        assert df.iloc[1]["found_by"] == "detect-secrets"

    def test_empty_data(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel([])
        assert os.path.exists(path)
        assert os.path.basename(path) == "myrepo_secrets_report.xlsx"
        df = pd.read_excel(path, sheet_name="General")
        assert len(df) == 0

    def test_tool_tab_content(self, tmp_path):
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        df_gl = pd.read_excel(path, sheet_name="Gitleaks")
        assert len(df_gl) == 1
        assert df_gl.iloc[0]["secret_value"] == "AKIA123"

    def test_auto_filters_applied(self, tmp_path):
        """Each sheet should have auto-filters enabled on all columns."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        wb = load_workbook(path)
        for ws in wb.worksheets:
            assert ws.auto_filter.ref == ws.dimensions, (
                f"Sheet '{ws.title}' missing auto-filter"
            )

    def test_auto_filters_on_empty(self, tmp_path):
        """Auto-filters should be present even on an empty report."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel([])
        wb = load_workbook(path)
        ws = wb["General"]
        assert ws.auto_filter.ref == ws.dimensions

    def test_repo_url_not_in_columns(self, tmp_path):
        """repo_url is internal metadata and must not appear as an Excel column."""
        reporter = Reporter(str(tmp_path), repo_name="myrepo")
        path = reporter.generate_excel(SAMPLE_DATA)
        df = pd.read_excel(path, sheet_name="General")
        assert "repo_url" not in df.columns
