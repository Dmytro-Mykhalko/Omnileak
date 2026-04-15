import pandas as pd
import pytest
from openpyxl import Workbook

from core.excel_utils import (
    sanitize_for_excel,
    SECRET_VALUE_LIMIT,
    ILLEGAL_CHARS_RE,
    build_commit_url,
    add_commit_hyperlinks,
)


class TestSanitizeForExcel:
    def test_strips_control_characters(self):
        df = pd.DataFrame([{"secret_value": "before\x00\x07\x0eafter", "other": "clean"}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["secret_value"] == "beforeafter"
        assert result.iloc[0]["other"] == "clean"

    def test_truncates_long_secret_value(self):
        long_val = "A" * (SECRET_VALUE_LIMIT + 500)
        df = pd.DataFrame([{"secret_value": long_val}])
        result = sanitize_for_excel(df)
        val = result.iloc[0]["secret_value"]
        assert val.endswith("[truncated]")
        assert len(val) == SECRET_VALUE_LIMIT + len(" [truncated]")

    def test_does_not_truncate_short_value(self):
        df = pd.DataFrame([{"secret_value": "short"}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["secret_value"] == "short"

    def test_does_not_modify_original(self):
        df = pd.DataFrame([{"secret_value": "before\x00after"}])
        sanitize_for_excel(df)
        assert df.iloc[0]["secret_value"] == "before\x00after"

    def test_handles_non_string_columns(self):
        df = pd.DataFrame([{"secret_value": "val", "line_number": 42}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["line_number"] == 42

    def test_handles_none_values(self):
        df = pd.DataFrame([{"secret_value": None, "other": None}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["secret_value"] is None

    def test_handles_empty_dataframe(self):
        df = pd.DataFrame(columns=["secret_value", "other"])
        result = sanitize_for_excel(df)
        assert len(result) == 0

    def test_works_without_secret_value_column(self):
        df = pd.DataFrame([{"name": "no\x00secret\x07here"}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["name"] == "nosecrethere"

    def test_truncation_at_exact_limit(self):
        exact = "B" * SECRET_VALUE_LIMIT
        df = pd.DataFrame([{"secret_value": exact}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["secret_value"] == exact  # not truncated

    def test_truncation_at_limit_plus_one(self):
        over = "C" * (SECRET_VALUE_LIMIT + 1)
        df = pd.DataFrame([{"secret_value": over}])
        result = sanitize_for_excel(df)
        assert result.iloc[0]["secret_value"].endswith("[truncated]")


class TestIllegalCharsRegex:
    def test_matches_null_byte(self):
        assert ILLEGAL_CHARS_RE.search("\x00")

    def test_matches_bell(self):
        assert ILLEGAL_CHARS_RE.search("\x07")

    def test_does_not_match_newline(self):
        assert not ILLEGAL_CHARS_RE.search("\n")

    def test_does_not_match_tab(self):
        assert not ILLEGAL_CHARS_RE.search("\t")

    def test_does_not_match_printable(self):
        assert not ILLEGAL_CHARS_RE.search("hello world 123 !@#")


class TestBuildCommitUrl:
    def test_github_url(self):
        url = build_commit_url("https://github.com/org/repo", "abc123", "src/app.py", 42)
        assert url == "https://github.com/org/repo/blob/abc123/src/app.py#L42"

    def test_gitlab_url(self):
        url = build_commit_url("https://gitlab.example.com/org/repo", "abc123", "src/app.py", 5)
        assert url == "https://gitlab.example.com/org/repo/-/blob/abc123/src/app.py#L5"

    def test_no_line_number(self):
        url = build_commit_url("https://github.com/org/repo", "abc123", "file.txt")
        assert url == "https://github.com/org/repo/blob/abc123/file.txt"

    def test_empty_repo_url(self):
        assert build_commit_url("", "abc123", "file.txt") == ""

    def test_empty_commit(self):
        assert build_commit_url("https://github.com/org/repo", "", "f.txt") == ""


class TestAddCommitHyperlinks:
    def _make_ws_with_data(self):
        """Create a minimal worksheet with a commit column and one data row."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="file_path")
        ws.cell(row=1, column=2, value="commit")
        ws.cell(row=1, column=3, value="line_number")
        ws.cell(row=2, column=1, value="src/app.py")
        ws.cell(row=2, column=2, value="abc1234")
        ws.cell(row=2, column=3, value=10)
        return ws

    def test_adds_hyperlink_from_dict_list(self):
        ws = self._make_ws_with_data()
        data = [{"commit": "abc1234", "file_path": "src/app.py", "line_number": 10}]
        add_commit_hyperlinks(ws, "https://github.com/org/repo", data)
        assert ws.cell(row=2, column=2).hyperlink is not None
        assert "abc1234" in ws.cell(row=2, column=2).hyperlink.target

    def test_adds_hyperlink_from_dataframe(self):
        ws = self._make_ws_with_data()
        df = pd.DataFrame([{"commit": "abc1234", "file_path": "src/app.py", "line_number": 10}])
        add_commit_hyperlinks(ws, "https://github.com/org/repo", df)
        assert ws.cell(row=2, column=2).hyperlink is not None

    def test_no_hyperlink_without_repo_url(self):
        ws = self._make_ws_with_data()
        data = [{"commit": "abc1234", "file_path": "src/app.py", "line_number": 10}]
        add_commit_hyperlinks(ws, "", data)
        assert ws.cell(row=2, column=2).hyperlink is None

    def test_no_crash_without_commit_column(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="other")
        add_commit_hyperlinks(ws, "https://github.com/org/repo", [{"other": "x"}])
