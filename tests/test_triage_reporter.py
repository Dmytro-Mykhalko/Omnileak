import json
import os

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from core.ai.triage_reporter import (
    convert,
    _flatten_lists,
    _sort_findings,
    FINDING_COLUMNS,
    COMPOSITE_COLUMNS,
)
from core.excel_utils import SECRET_VALUE_LIMIT


# ── Fixtures (synthetic only — no real secrets) ─────────────────────────────

def _make_tp(id_=1, severity="HIGH", on_disk=True, environment="production"):
    return {
        "id": id_,
        "omnileak_ids": [id_, id_ + 100],
        "classification": "TRUE_POSITIVE",
        "severity": severity,
        "category": "Generic API Key",
        "secret_value": f"sk_test_{id_}_FAKEFAKEFAKE",
        "file_path": f"config/app_{id_}.yml",
        "line_number": id_ * 10,
        "commit": f"{'a' * 7}{id_}",
        "on_disk": on_disk,
        "confidence": "high",
        "environment": environment,
        "remediation": "ROTATE_IMMEDIATELY",
        "effort": "quick",
        "detected_by": ["gitleaks", "trufflehog"],
        "fp_reason": None,
        "duplicate_of": None,
    }


def _make_dup(id_=50, duplicate_of=1):
    return {
        "id": id_,
        "omnileak_ids": [id_],
        "classification": "DUPLICATE",
        "severity": "HIGH",
        "category": "Generic API Key",
        "secret_value": "sk_test_1_FAKEFAKEFAKE",
        "file_path": "config/app_1.yml",
        "line_number": 10,
        "commit": f"{'c' * 7}{id_}",
        "on_disk": False,
        "confidence": None,
        "environment": None,
        "remediation": None,
        "effort": None,
        "detected_by": ["gitleaks"],
        "fp_reason": None,
        "duplicate_of": duplicate_of,
    }


def _make_fp(id_=99):
    return {
        "id": id_,
        "omnileak_ids": [id_],
        "classification": "FALSE_POSITIVE",
        "severity": None,
        "category": "vendor translation",
        "secret_value": "password => 'Jelszó'",
        "file_path": "vendor/lang/hu.php",
        "line_number": 42,
        "commit": f"{'b' * 7}{id_}",
        "on_disk": True,
        "confidence": None,
        "environment": None,
        "remediation": None,
        "effort": None,
        "detected_by": ["detect-secrets"],
        "fp_reason": "Translation file in vendor directory",
        "duplicate_of": None,
    }


def _make_composite():
    return {
        "id": 1,
        "description": "Private key and passphrase both committed",
        "severity": "CRITICAL",
        "related_finding_ids": [1, 2],
        "files_involved": ["config/key.pem", "config/passphrase.txt"],
    }


def _make_meta(**overrides):
    meta = {
        "repo": "test-repo",
        "scan_date": "2026-04-15T10:00:00Z",
        "last_commit": "abc1234",
        "mode": "full_scan",
        "risk_score": 45,
        "total_raw_findings": 10,
        "true_positives": 3,
        "false_positives_filtered": 7,
        "ai_only_findings": 1,
        "deep_analysis_performed": True,
    }
    meta.update(overrides)
    return meta


def _write_json(tmp_path, findings=None, composites=None, meta=None):
    """Write a triage JSON file and return its path."""
    data = {
        "meta": meta or _make_meta(),
        "findings": findings if findings is not None else [_make_tp(), _make_fp()],
        "composite_vulnerabilities": composites or [],
    }
    path = tmp_path / "triage-results.json"
    path.write_text(json.dumps(data), encoding="utf-8")
    return str(path)


# ── convert() — sheet structure ─────────────────────────────────────────────

class TestSheetStructure:
    def test_creates_xlsx(self, tmp_path):
        json_path = _write_json(tmp_path)
        xlsx = convert(json_path)
        assert os.path.isfile(xlsx)
        assert xlsx.endswith(".xlsx")

    def test_default_output_name(self, tmp_path):
        json_path = _write_json(tmp_path)
        xlsx = convert(json_path)
        assert os.path.basename(xlsx) == "triage-results.xlsx"
        assert os.path.dirname(xlsx) == str(tmp_path)

    def test_custom_output_path(self, tmp_path):
        json_path = _write_json(tmp_path)
        custom = str(tmp_path / "custom_report.xlsx")
        xlsx = convert(json_path, excel_path=custom)
        assert xlsx == custom
        assert os.path.isfile(custom)

    def test_has_expected_sheets(self, tmp_path):
        json_path = _write_json(
            tmp_path,
            findings=[_make_tp(), _make_fp()],
            composites=[_make_composite()],
        )
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "Summary" in wb.sheetnames
        assert "All Findings" in wb.sheetnames
        assert "True Positives" in wb.sheetnames
        assert "False Positives" in wb.sheetnames
        assert "Composite Vulns" in wb.sheetnames

    def test_no_tp_sheet_when_all_fps(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_fp(1), _make_fp(2)])
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "True Positives" not in wb.sheetnames
        assert "False Positives" in wb.sheetnames

    def test_no_fp_sheet_when_all_tps(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp(1), _make_tp(2)])
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "True Positives" in wb.sheetnames
        assert "False Positives" not in wb.sheetnames

    def test_no_composite_sheet_when_empty(self, tmp_path):
        json_path = _write_json(tmp_path, composites=[])
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "Composite Vulns" not in wb.sheetnames


# ── convert() — empty / edge cases ──────────────────────────────────────────

class TestEdgeCases:
    def test_empty_findings(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[], composites=[])
        xlsx = convert(json_path)
        assert os.path.isfile(xlsx)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert len(df) == 0

    def test_empty_meta(self, tmp_path):
        json_path = _write_json(tmp_path, meta={})
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Summary")
        assert len(df) == 10  # all metric rows present, values empty


# ── Summary sheet ────────────────────────────────────────────────────────────

class TestSummarySheet:
    def test_contains_meta_fields(self, tmp_path):
        meta = _make_meta(risk_score=72)
        json_path = _write_json(tmp_path, meta=meta)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Summary")
        metrics = dict(zip(df["Metric"], df["Value"]))
        assert metrics["Repository"] == "test-repo"
        assert metrics["Risk Score"] == 72
        assert metrics["Mode"] == "full_scan"

    def test_bold_headers(self, tmp_path):
        json_path = _write_json(tmp_path)
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        ws = wb["Summary"]
        for col in range(1, ws.max_column + 1):
            assert ws.cell(row=1, column=col).font.bold


# ── All Findings sheet — data correctness ────────────────────────────────────

class TestAllFindings:
    def test_row_count(self, tmp_path):
        findings = [_make_tp(1), _make_tp(2), _make_fp(3)]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert len(df) == 3

    def test_tps_sorted_before_fps(self, tmp_path):
        findings = [_make_fp(1), _make_tp(2)]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert df.iloc[0]["classification"] == "TRUE_POSITIVE"
        assert df.iloc[1]["classification"] == "FALSE_POSITIVE"

    def test_severity_sort_within_tps(self, tmp_path):
        findings = [
            _make_tp(1, severity="LOW"),
            _make_tp(2, severity="CRITICAL"),
            _make_tp(3, severity="MEDIUM"),
        ]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert list(df["severity"]) == ["CRITICAL", "MEDIUM", "LOW"]

    def test_detected_by_flattened(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp()])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert df.iloc[0]["detected_by"] == "gitleaks, trufflehog"

    def test_omnileak_ids_flattened(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp(5)])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert df.iloc[0]["omnileak_ids"] == "5, 105"

    def test_fp_reason_present_for_fps(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_fp()])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert "vendor" in str(df.iloc[0]["fp_reason"]).lower()

    def test_column_order(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp()])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert list(df.columns) == FINDING_COLUMNS


# ── Formatting ───────────────────────────────────────────────────────────────

class TestFormatting:
    def test_auto_filters_on_all_sheets(self, tmp_path):
        json_path = _write_json(
            tmp_path,
            findings=[_make_tp(), _make_fp()],
            composites=[_make_composite()],
        )
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        for ws in wb.worksheets:
            if ws.title != "Summary":
                assert ws.auto_filter.ref == ws.dimensions, (
                    f"Sheet '{ws.title}' missing auto-filter"
                )

    def test_bold_headers_on_finding_sheets(self, tmp_path):
        json_path = _write_json(tmp_path)
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        ws = wb["All Findings"]
        for col in range(1, ws.max_column + 1):
            assert ws.cell(row=1, column=col).font.bold

    def test_severity_cells_colored(self, tmp_path):
        findings = [
            _make_tp(1, severity="CRITICAL"),
            _make_tp(2, severity="HIGH"),
            _make_tp(3, severity="MEDIUM"),
            _make_tp(4, severity="LOW"),
        ]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        ws = wb["All Findings"]

        sev_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "severity":
                sev_col = col
                break
        assert sev_col is not None

        # Each severity row should have a non-default fill
        for row in range(2, 6):
            fill = ws.cell(row=row, column=sev_col).fill
            assert fill.fill_type == "solid"

    def test_fp_classification_grayed(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp(1), _make_fp(2)])
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        ws = wb["All Findings"]

        cls_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "classification":
                cls_col = col
                break

        # Find the FP row (sorted: TP first, FP second)
        fp_row = 3  # row 1=header, row 2=TP, row 3=FP
        fill = ws.cell(row=fp_row, column=cls_col).fill
        assert fill.start_color.rgb == "00D3D3D3"


# ── Sanitization ─────────────────────────────────────────────────────────────

class TestSanitization:
    def test_control_chars_stripped(self, tmp_path):
        tp = _make_tp()
        tp["secret_value"] = "before\x00\x07\x0eafter"
        json_path = _write_json(tmp_path, findings=[tp])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert df.iloc[0]["secret_value"] == "beforeafter"

    def test_long_secret_truncated(self, tmp_path):
        tp = _make_tp()
        tp["secret_value"] = "A" * (SECRET_VALUE_LIMIT + 500)
        json_path = _write_json(tmp_path, findings=[tp])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        val = df.iloc[0]["secret_value"]
        assert len(val) <= SECRET_VALUE_LIMIT + 20  # room for " [truncated]"
        assert val.endswith("[truncated]")

    def test_short_secret_not_truncated(self, tmp_path):
        tp = _make_tp()
        tp["secret_value"] = "short_value"
        json_path = _write_json(tmp_path, findings=[tp])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert df.iloc[0]["secret_value"] == "short_value"


# ── Composite vulnerabilities sheet ──────────────────────────────────────────

class TestCompositeVulns:
    def test_composite_row_count(self, tmp_path):
        json_path = _write_json(
            tmp_path,
            composites=[_make_composite(), _make_composite()],
        )
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Composite Vulns")
        assert len(df) == 2

    def test_composite_columns(self, tmp_path):
        json_path = _write_json(tmp_path, composites=[_make_composite()])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Composite Vulns")
        assert list(df.columns) == COMPOSITE_COLUMNS

    def test_composite_lists_flattened(self, tmp_path):
        json_path = _write_json(tmp_path, composites=[_make_composite()])
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Composite Vulns")
        assert df.iloc[0]["related_finding_ids"] == "1, 2"
        assert "key.pem" in df.iloc[0]["files_involved"]


# ── Internal helpers ─────────────────────────────────────────────────────────

class TestSortFindings:
    def test_tps_before_fps(self):
        df = pd.DataFrame([
            {"id": 1, "classification": "FALSE_POSITIVE", "severity": None},
            {"id": 2, "classification": "TRUE_POSITIVE", "severity": "HIGH"},
        ])
        result = _sort_findings(df)
        assert result.iloc[0]["classification"] == "TRUE_POSITIVE"

    def test_critical_before_low(self):
        df = pd.DataFrame([
            {"id": 1, "classification": "TRUE_POSITIVE", "severity": "LOW"},
            {"id": 2, "classification": "TRUE_POSITIVE", "severity": "CRITICAL"},
        ])
        result = _sort_findings(df)
        assert result.iloc[0]["severity"] == "CRITICAL"


class TestFlattenLists:
    def test_list_to_string(self):
        df = pd.DataFrame([{"detected_by": ["a", "b"], "omnileak_ids": [1, 2]}])
        result = _flatten_lists(df.copy())
        assert result.iloc[0]["detected_by"] == "a, b"
        assert result.iloc[0]["omnileak_ids"] == "1, 2"

    def test_non_list_unchanged(self):
        df = pd.DataFrame([{"detected_by": "already_string"}])
        result = _flatten_lists(df.copy())
        assert result.iloc[0]["detected_by"] == "already_string"


# ── Duplicate classification ─────────────────────────────────────────────────

class TestDuplicates:
    def test_dup_sheet_created(self, tmp_path):
        findings = [_make_tp(1), _make_dup(2, duplicate_of=1)]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "Duplicates" in wb.sheetnames

    def test_no_dup_sheet_when_none(self, tmp_path):
        json_path = _write_json(tmp_path, findings=[_make_tp(1)])
        xlsx = convert(json_path)
        wb = load_workbook(xlsx)
        assert "Duplicates" not in wb.sheetnames

    def test_dup_sort_order(self):
        """DUPs should sort between TPs and FPs."""
        df = pd.DataFrame([
            {"id": 1, "classification": "FALSE_POSITIVE", "severity": None},
            {"id": 2, "classification": "DUPLICATE", "severity": "HIGH"},
            {"id": 3, "classification": "TRUE_POSITIVE", "severity": "HIGH"},
        ])
        result = _sort_findings(df)
        assert list(result["classification"]) == [
            "TRUE_POSITIVE", "DUPLICATE", "FALSE_POSITIVE"
        ]

    def test_dup_in_all_findings(self, tmp_path):
        findings = [_make_tp(1), _make_dup(2, duplicate_of=1), _make_fp(3)]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="All Findings")
        assert len(df) == 3
        assert list(df["classification"]) == [
            "TRUE_POSITIVE", "DUPLICATE", "FALSE_POSITIVE"
        ]

    def test_dup_has_duplicate_of_field(self, tmp_path):
        findings = [_make_tp(1), _make_dup(2, duplicate_of=1)]
        json_path = _write_json(tmp_path, findings=findings)
        xlsx = convert(json_path)
        df = pd.read_excel(xlsx, sheet_name="Duplicates")
        assert df.iloc[0]["duplicate_of"] == 1


# ── File naming ──────────────────────────────────────────────────────────────

class TestFileNaming:
    def test_xlsx_mirrors_json_name(self, tmp_path):
        """When json is named repo_triage-results_72.json, xlsx should match."""
        data = {
            "meta": _make_meta(),
            "findings": [_make_tp()],
            "composite_vulnerabilities": [],
        }
        json_path = str(tmp_path / "myapp_triage-results_72.json")
        with open(json_path, "w") as f:
            json.dump(data, f)
        xlsx = convert(json_path)
        assert os.path.basename(xlsx) == "myapp_triage-results_72.xlsx"

    def test_default_name_for_plain_json(self, tmp_path):
        json_path = _write_json(tmp_path)  # writes triage-results.json
        xlsx = convert(json_path)
        assert os.path.basename(xlsx) == "triage-results.xlsx"
